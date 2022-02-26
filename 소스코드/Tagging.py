import openpyxl as xl
import re
import pandas as pd
import Function

"""
태그 목록
------------------------------------
<F> Feature 태그
<PB> 전치 서술어(형용사) 태그
<PA> 후치 서술어 태그
<NB> 전치 부정어 태그
<NA> 후치 부정어 태그
<EB> 전치 강조 태그
<EA> 후치 강조 태그
------------------------------------
"""

#리스트
df_list = ['문장', '수식관계/평점', 'Unnamed: 6', 'Unnamed: 7', 'Unnamed: 8']
str_list = ['','','','','']
tag_list = [['<F>', '</F>'],['PB','PA'],['NB','NA'],['EB','EA']]

#태깅할 파일 제목 확장자 포함해서 입력
filename_for_tagging = '서울관광지.xlsx'

filename_for_saving = filename_for_tagging[:-5] + ' 태깅.txt'
saving = open(filename_for_saving, 'w', encoding='UTF-8')

filename_for_errlog = filename_for_tagging[:-5] + ' 에러로그.txt'
errlog = open(filename_for_errlog, 'w')

wb = xl.load_workbook('./' + filename_for_tagging)

###Code
for i in range(0, len(wb.sheetnames)):

    if i == 0:
        saving_str = "[" + wb.sheetnames[i] + "]"
    else:
        saving_str = saving_str + "\n[" + wb.sheetnames[i] + "]"

    df = pd.read_excel('./' + filename_for_tagging, sheet_name=wb.sheetnames[i])

    for j in df_list:
        df[j] = df[j].str.strip().copy()
        df[j] = df[j].str.strip('\n').copy()
        
    df = df.fillna(pd.NA).copy()

    for j in range(1, df.shape[0]):
        
        for k in range(len(df_list)):
            str_list[k] = df[df_list[k]][j]
            
        if str_list[2] is not pd.NA:
            Function.tagging(1, df_list, str_list, tag_list, j, df, errlog)
            
            if str_list[3] is not pd.NA:
                Function.tagging(2, df_list, str_list, tag_list, j, df, errlog)
                
            if str_list[4] is not pd.NA:
                Function.tagging(3, df_list, str_list, tag_list, j, df, errlog)
                
        elif str_list[1] is pd.NA:
            continue

        if str_list[0] is not pd.NA:
            saving_str = saving_str + "\n" + str_list[0]
                
saving.write(saving_str)
